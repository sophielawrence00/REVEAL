"""
Merges REVEAL's Pearson r results (per protein CSV) with the SNR data
(snr_long.csv) and builds an r-vs-SNR scatter plot per protein, with
threshold lines marking the "confident binder" quadrant
(r > threshold AND SNR > SNR_CUTOFF).
"""
import os
import glob
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT_DIR   = r"DIA_DIR"   # <-- update if needed
RESULTS_DIR = os.path.join(ROOT_DIR, "REVEAL_correlation_output")
SNR_CSV     = os.path.join(ROOT_DIR, "snr_long.csv")
SNR_CUTOFF  = 3.0

snr = pd.read_csv(SNR_CSV)

result_files = glob.glob(os.path.join(RESULTS_DIR, "*_pearson_results.csv"))
if not result_files:
    raise FileNotFoundError(f"No *_pearson_results.csv files found in {RESULTS_DIR}")

confident_binders_by_protein = {}   # protein -> DataFrame, for the Excel export
apo_by_protein = {}                 # protein -> merged apo-condition rows, for the comparison plots

for fp in result_files:
    protein = os.path.basename(fp).replace("_pearson_results.csv", "")
    df = pd.read_csv(fp)
    id_cols = ["mz_key", "lipid_primary", "lipid_candidates",
               "classified_binder_any", "threshold_r"]
    cond_cols = [c for c in df.columns if c not in id_cols]

    long_r = df.melt(id_vars=["mz_key", "lipid_primary", "threshold_r"], value_vars=cond_cols,
                      var_name="condition", value_name="pearson_r").dropna(subset=["pearson_r"])

    protein_snr = snr[snr["protein"] == protein]
    merged = long_r.merge(protein_snr, on=["mz_key", "condition"], how="inner")

    # Diagnostic: show exactly what condition names exist on each side
    r_conditions   = set(long_r["condition"].unique())
    snr_conditions = set(protein_snr["condition"].unique())
    print(f"  {protein}: conditions in Pearson results = {sorted(r_conditions)}")
    print(f"  {protein}: conditions in SNR file        = {sorted(snr_conditions)}")
    if r_conditions - snr_conditions:
        print(f"  {protein}: MISSING from SNR file: {sorted(r_conditions - snr_conditions)}")
    if snr_conditions - r_conditions:
        print(f"  {protein}: MISSING from Pearson results (unexpected extra names in SNR file): {sorted(snr_conditions - r_conditions)}")

    if merged.empty:
        print(f"  WARNING: no overlapping mz_key/condition rows for {protein} — skipping plot")
        continue

    r_threshold = merged["threshold_r"].iloc[0]
    merged["confident_binder"] = (merged["pearson_r"] > r_threshold) & (merged["snr_mean"] > SNR_CUTOFF)

    # --- Pass/fail summary ---
    total_tested = len(merged)
    total_passed = merged["confident_binder"].sum()
    total_failed = total_tested - total_passed
    print(f"  {protein}: {total_passed} passed / {total_failed} failed  ({total_tested} lipid-condition combinations tested)")

    per_cond = merged.groupby("condition")["confident_binder"].agg(
        passed="sum", tested="count"
    )
    per_cond["failed"] = per_cond["tested"] - per_cond["passed"]
    for cond, row in per_cond.iterrows():
        print(f"    {cond:<12} passed: {row['passed']:>3}   failed: {row['failed']:>3}   tested: {row['tested']:>3}")

    fig, ax = plt.subplots(figsize=(7, 6))
    conditions = merged["condition"].unique()
    colors = plt.cm.tab10.colors
    for i, cond in enumerate(conditions):
        sub = merged[merged["condition"] == cond]
        ax.scatter(sub["pearson_r"], sub["snr_mean"],
                   label=cond, color=colors[i % len(colors)],
                   edgecolor="black", linewidth=0.4, s=45, alpha=0.85)

    ax.axvline(r_threshold, color="grey", linestyle="--", linewidth=1)
    ax.axhline(SNR_CUTOFF, color="grey", linestyle="--", linewidth=1)

    ax.set_xlabel("Pearson r (vs MS1 envelope)")
    ax.set_ylabel("Signal-to-noise ratio")
    ax.set_title(f"{protein} — Pearson r vs SNR\n"
                 f"(confident binder = r > {r_threshold:.3f} AND SNR > {SNR_CUTOFF:.1f})")
    ax.legend(title="Condition", fontsize=8)
    fig.tight_layout()

    out_path = os.path.join(RESULTS_DIR, f"{protein}_r_vs_snr_scatter.pdf")
    fig.savefig(out_path, dpi=300)
    plt.close(fig)
    print(f"  Saved: {out_path}  ({merged['confident_binder'].sum()} confident binders out of {len(merged)} points)")

    # --- Bar chart: binder counts per condition, using BOTH criteria ---
    counts = merged.groupby("condition")["confident_binder"].sum()
    counts = counts.reindex(cond_cols).dropna()  # keep original condition order where possible

    fig2, ax2 = plt.subplots(figsize=(max(4, len(counts) * 1.2), 3.5))
    bars = ax2.bar(counts.index, counts.values,
                   color="#2a7f62", edgecolor="white", linewidth=0.5)
    ax2.bar_label(bars, fontsize=9)
    ax2.set_ylabel("Confident binders", fontsize=10)
    ax2.set_title(f"{protein}  —  Binders per condition\n"
                  f"(r > {r_threshold:.3f} AND SNR > {SNR_CUTOFF:.1f})",
                  fontsize=10, fontweight="bold")
    ax2.set_ylim(0, max(counts.values.max(), 1) * 1.35)
    for spine in ["top", "right"]:
        ax2.spines[spine].set_visible(False)
    fig2.tight_layout()
    out_path2 = os.path.join(RESULTS_DIR, f"{protein}_binder_counts_dual.pdf")
    fig2.savefig(out_path2, dpi=300, bbox_inches="tight")
    plt.close(fig2)
    print(f"  Saved: {out_path2}")

    # --- Collect confident binders for the Excel workbook ---
    binders = merged[merged["confident_binder"]][
        ["mz_key", "lipid_primary", "condition", "pearson_r", "snr_mean"]
    ].sort_values(["condition", "mz_key"]).reset_index(drop=True)
    confident_binders_by_protein[protein] = binders

    # --- Stash this protein's apo-condition rows for the cross-protein comparison below ---
    apo_rows = merged[merged["condition"] == "apo"].copy()
    if not apo_rows.empty:
        apo_rows["protein"] = protein
        apo_by_protein[protein] = apo_rows

# ═══════════════════════════════════════════════════════════════════════════
# mGlyR apo vs mGluR2 apo — direct comparison plots
# ═══════════════════════════════════════════════════════════════════════════
if "mGlyR" in apo_by_protein and "mGluR2" in apo_by_protein:
    apo_combined = pd.concat(
        [apo_by_protein["mGlyR"], apo_by_protein["mGluR2"]], ignore_index=True
    )

    # --- Comparison scatter: r vs SNR, colored by protein ---
    fig3, ax3 = plt.subplots(figsize=(7, 6))
    protein_colors = {"mGlyR": "#d62728", "mGluR2": "#1f77b4"}
    for protein in ["mGlyR", "mGluR2"]:
        sub = apo_combined[apo_combined["protein"] == protein]
        ax3.scatter(sub["pearson_r"], sub["snr_mean"],
                    label=f"{protein} (apo)", color=protein_colors[protein],
                    edgecolor="black", linewidth=0.4, s=45, alpha=0.85)
        # each protein has its own threshold — draw a matching vertical line
        r_thresh = sub["threshold_r"].iloc[0]
        ax3.axvline(r_thresh, color=protein_colors[protein], linestyle="--", linewidth=1,
                    label=f"{protein} threshold ({r_thresh:.2f})")

    ax3.axhline(SNR_CUTOFF, color="grey", linestyle="--", linewidth=1, label=f"SNR = {SNR_CUTOFF:.1f}")
    ax3.set_xlabel("Pearson r (vs MS1 envelope)")
    ax3.set_ylabel("Signal-to-noise ratio")
    ax3.set_title("mGlyR apo vs mGluR2 apo — Pearson r vs SNR")
    ax3.legend(fontsize=8)
    fig3.tight_layout()
    out_path3 = os.path.join(RESULTS_DIR, "mGlyR_vs_mGluR2_apo_scatter.pdf")
    fig3.savefig(out_path3, dpi=300)
    plt.close(fig3)
    print(f"  Saved: {out_path3}")

    # --- Comparison bar chart: confident binder counts, apo only ---
    apo_counts = apo_combined.groupby("protein")["confident_binder"].sum().reindex(["mGlyR", "mGluR2"])
    fig4, ax4 = plt.subplots(figsize=(4.5, 3.5))
    bars4 = ax4.bar(apo_counts.index, apo_counts.values,
                     color=[protein_colors[p] for p in apo_counts.index],
                     edgecolor="white", linewidth=0.5)
    ax4.bar_label(bars4, fontsize=9)
    ax4.set_ylabel("Confident binders (apo)", fontsize=10)
    ax4.set_title("mGlyR apo vs mGluR2 apo\n(r > protein threshold AND SNR > 3.0)",
                  fontsize=10, fontweight="bold")
    ax4.set_ylim(0, max(apo_counts.values.max(), 1) * 1.35)
    for spine in ["top", "right"]:
        ax4.spines[spine].set_visible(False)
    fig4.tight_layout()
    out_path4 = os.path.join(RESULTS_DIR, "mGlyR_vs_mGluR2_apo_binder_counts.pdf")
    fig4.savefig(out_path4, dpi=300, bbox_inches="tight")
    plt.close(fig4)
    print(f"  Saved: {out_path4}")
else:
    print("  Skipping mGlyR vs mGluR2 apo comparison — apo data not available for both proteins yet")

# ═══════════════════════════════════════════════════════════════════════════
# Write confident binders (r > threshold AND SNR > cutoff) to one Excel
# workbook, one sheet per protein
# ═══════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2A7F62")
body_font   = Font(name="Arial")

for protein, binders in confident_binders_by_protein.items():
    ws = wb.create_sheet(title=protein[:31])  # Excel sheet name limit
    headers = ["m/z", "Lipid", "Condition", "Pearson r", "SNR"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for _, row in binders.iterrows():
        ws.append([row["mz_key"], row["lipid_primary"], row["condition"],
                   round(row["pearson_r"], 3), round(row["snr_mean"], 2)])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    # Auto-width columns
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(10, length + 2)

if len(wb.sheetnames) == 0:
    wb.create_sheet(title="No data")

xlsx_path = os.path.join(RESULTS_DIR, "confident_binders.xlsx")
wb.save(xlsx_path)
print(f"\nSaved confident binders workbook: {xlsx_path}")
